"""
Message Portfolio Pipeline
=========================
接收用户文本持仓数据 → 解析保存为 Excel → 复用 data-pipeline 标准化流程入库 MongoDB。

相比 Image OCR Pipeline：用户直接提供文本/TSV/CSV，无需 OCR 识别。

Usage:
  python scripts/run_string_pipeline.py -i "raw text..." -d 2026-05-03
  python scripts/run_string_pipeline.py -f /path/to/input.txt -d 2026-05-03
"""
import argparse
import io
import re
import sys
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers

# 复用 data-pipeline 模块
_pipeline_root = Path(__file__).resolve().parents[2] / "data-pipeline" / "scripts"
sys.path.insert(0, str(_pipeline_root))

from transformers.portfolio_excel_transformer import PortfolioExcelTransformer
from transformers.image_portfolio_normalizer import normalize_all
from validators.portfolio_validator import (
    validate_basic_info, validate_nav, validate_position,
    ValidationResult,
)
from loaders.mongodb_loader import PortfolioMongoLoader


STANDARD_COLUMNS = [
    "截止日期", "产品名称", "产品代码", "Wind代码", "资产名称",
    "持仓比例", "数量", "市值（本币）", "最新净值", "最新份额", "最新规模"
]

COLUMN_ALIASES = {
    "截止日期": ["截止日期", "日期", "date"],
    "产品名称": ["产品名称", "产品", "name"],
    "产品代码": ["产品代码", "代码", "code"],
    "Wind代码": ["Wind代码", "wind_code", "wind", "Wind", "证券代码"],
    "资产名称": ["资产名称", "资产", "asset", "股票名称"],
    "持仓比例": ["持仓比例", "比例", "ratio", "持有比例"],
    "数量": ["数量", "qty", "shares", "股数"],
    "市值（本币）": ["市值（本币）", "市值(本币)", "市值", "mkt", "market_value"],
    "最新净值": ["最新净值", "净值", "nav"],
    "最新份额": ["最新份额", "份额", "share"],
    "最新规模": ["最新规模", "规模", "aum"],
}


def normalize_column(col: str) -> Optional[str]:
    col_clean = col.strip()
    for std_name, aliases in COLUMN_ALIASES.items():
        if col_clean in aliases or col_clean == std_name:
            return std_name
    return None


def detect_separator(text: str) -> str:
    lines = text.strip().splitlines()
    if not lines:
        return "\t"
    first_line = lines[0]
    if "\t" in first_line:
        return "\t"
    if "," in first_line:
        return ","
    if "|" in first_line:
        return "|"
    return "fixed"


def parse_text_to_df(text: str) -> pd.DataFrame:
    """将原始文本解析为 pandas DataFrame。"""
    text = text.strip()
    if not text:
        raise ValueError("输入文本为空")

    sep = detect_separator(text)

    if sep != "fixed":
        df = pd.read_csv(
            io.StringIO(text), sep=sep, header=0,
            dtype=str, keep_default_na=False,
        )
        df = df.dropna(how="all")
        df = df.rename(columns={c: c.strip() for c in df.columns})
    else:
        raise ValueError("固定宽度格式暂不支持，请使用 Tab 或逗号分隔")

    # 列名映射
    renamed = {}
    for col in df.columns:
        std = normalize_column(col)
        if std:
            renamed[col] = std
    df = df.rename(columns=renamed)

    # 补全标准列
    for col in STANDARD_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[STANDARD_COLUMNS]

    # 数值列（持仓比例需要转换为小数）
    for col in ["持仓比例", "数量", "市值（本币）", "最新净值", "最新份额", "最新规模"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].astype(str).str.strip().str.replace("%", "", regex=False).str.replace("％", "", regex=False), errors="coerce")
            if col == "持仓比例":
                df[col] = df[col] / 100  # Convert percentage to decimal

    # 字符串列
    for col in ["截止日期", "产品名称", "产品代码", "Wind代码", "资产名称"]:
        df[col] = df[col].astype(str).str.strip().replace(["", "nan", "None"], "")

    df = df[df["产品代码"].notna() & (df["产品代码"] != "")]
    return df


def save_raw_txt(raw_text: str, date_str: str, source_root: Path) -> Path:
    """
    保存原始消息文本到 source/smart-money/{date}/message/。
    文件名格式：portfolio_YYYYMMDD_HHMM.txt（HHMM 为系统收到消息的时间，4位）
    """
    from datetime import datetime

    message_dir = source_root / date_str / "message"
    message_dir.mkdir(parents=True, exist_ok=True)

    receive_time = datetime.now()
    date_clean = receive_time.strftime("%Y%m%d")      # YYYYMMDD
    time_clean = receive_time.strftime("%H%M%S")         # HHMMSS，6位精确到秒
    base = f"portfolio_{date_clean}_{time_clean}"

    txt_path = message_dir / f"{base}.txt"
    counter = 1
    while txt_path.exists():
        txt_path = message_dir / f"{base}_{counter:02d}.txt"
        counter += 1

    txt_path.write_text(raw_text, encoding="utf-8")
    return txt_path


def save_excel(df: pd.DataFrame, date_str: str, source_root: Path, base_name: str) -> Path:
    """
    保存 Excel 到 source/smart-money/{date}/message/。
    文件名与同时间生成的 raw txt 文件名一致（仅扩展名不同）。

    Args:
        base_name: 对应的 raw txt 文件名（不含扩展名），
                   由 save_raw_txt 根据收到时间生成。
    """
    message_dir = source_root / date_str / "message"
    message_dir.mkdir(parents=True, exist_ok=True)

    excel_path = message_dir / f"{base_name}.xlsx"
    counter = 1
    while excel_path.exists():
        excel_path = message_dir / f"{base_name}_{counter:02d}.xlsx"
        counter += 1

    df.to_excel(excel_path, index=False, sheet_name="Sheet1")

    # 应用 Excel 格式：持仓比例列设为百分比格式（保留2位小数）
    # 如果是数值（如 5.77），需要除以 100 后 Excel 才能正确显示为 5.77%
    wb = load_workbook(excel_path)
    ws = wb.active

    # 找到持仓比例列的索引（0-based，Excel列从1开始）
    header = [cell.value for cell in ws[1]]
    if "持仓比例" in header:
        col_idx = header.index("持仓比例") + 1  # openpyxl列从1开始
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                cell.value = cell.value / 100  # 转为小数（如 0.0577）
                cell.number_format = '0.00%'   # Excel 百分比格式

    wb.save(excel_path)
    return excel_path


async def run_pipeline(
    raw_text: str,
    date_str: str,
    source_root: Path,
    dry_run: bool = False,
) -> dict:
    """
    完整流程：消息文本 → Excel → normalize → validate → MongoDB

    直接复用 data-pipeline 的 transformer，不经过 OCR。
    """
    # Step 1: 解析文本 → DataFrame → Excel
    df = parse_text_to_df(raw_text)
    if df.empty:
        raise ValueError("解析后无有效数据")

    # Step 1: 保存原始消息 + 生成 Excel
    # 收到时间作为文件名时间戳（区分多批次消息）
    txt_path = save_raw_txt(raw_text, date_str, source_root)
    base_name = txt_path.stem  # portfolio_YYYYMMDD_HHMM 或 portfolio_YYYYMMDD_HHMM_NN
    excel_path = save_excel(df, date_str, source_root, base_name)
    print(f"[Step1] Raw TXT → {txt_path}")
    print(f"[Step1] Excel  → {excel_path}  ({len(df)} rows)")

    if dry_run:
        return {"excel_path": str(excel_path), "rows": len(df), "dry_run": True}

    # Step 2: 读取 Excel → nested JSON（直接读取，不走 OCR）
    records = [{"df": df}]
    transformer = PortfolioExcelTransformer()
    nested = transformer.transform(records)
    normalized = normalize_all(nested[0])

    print(f"[Step2] Normalized: basic_info={len(normalized['basic_info'])}, "
          f"nav={len(normalized['nav'])}, position={len(normalized['position'])}")

    # Step 3: validate
    vr_basic = validate_basic_info(normalized["basic_info"])
    vr_nav = validate_nav(normalized["nav"])
    vr_pos = validate_position(normalized["position"])
    merged = ValidationResult()
    merged.merge(vr_basic)
    merged.merge(vr_nav)
    merged.merge(vr_pos)

    if merged.has_errors:
        print(f"[Step3] Validation errors: {merged.errors}")
        raise ValueError(f"Validation failed: {'; '.join(merged.errors)}")

    # Step 4: MongoDB
    loader = PortfolioMongoLoader()
    result = loader.load_all(normalized)

    print(f'[Step4] MongoDB: basic_info={result["basic_info"]}, nav={result["nav"]}, position={result["position"]}')

    return {
        "excel_path": str(excel_path),
        "rows": len(df),
        "basic_info": len(normalized["basic_info"]),
        "nav": len(normalized["nav"]),
        "position": len(normalized["position"]),
        "validation": {
            "valid": merged.valid,
            "errors": merged.errors,
            "warnings": merged.warnings,
        },
        "mongodb": {
            "inserted": result.get("inserted", 0),
            "updated": result.get("updated", 0),
            "skipped": result.get("skipped", 0),
        },
    }


def main():
    parser = argparse.ArgumentParser(description="Message Portfolio Pipeline")
    parser.add_argument("-i", "--input", help="原始文本")
    parser.add_argument("-f", "--file", help="从文件读取")
    parser.add_argument("-d", "--date", required=True, help="日期，如 2026-05-03")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if args.input:
        raw_text = args.input
    elif args.file:
        raw_text = Path(args.file).read_text(encoding="utf-8")
    else:
        print("请使用 -i '文本' 或 -f /path/to/file.txt 提供数据")
        sys.exit(1)

    # source_root: 向上两级到 workspace-yquant/skills/data/source/smart-money
    source_root = Path(__file__).resolve().parents[2] / "source" / "smart-money"

    import asyncio
    result = asyncio.run(run_pipeline(raw_text, args.date, source_root, dry_run=args.dry_run))
    print(f"\n✅ 完成：{result}")


if __name__ == "__main__":
    main()
