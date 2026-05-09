"""
Image Portfolio Pipeline
=======================
接收图片持仓数据 → MiniMax OCR → 保存 Excel → 复用 data-pipeline 标准化流程入库 MongoDB。

相比 Message Pipeline：图片需要经过 VLM OCR 识别，不能直接解析文本。

Usage:
    python scripts/run_image_pipeline.py -i /path/to/image.jpg -d 2026-05-09
    python scripts/run_image_pipeline.py -i /path/to/image.jpg -d 2026-05-09 --dry-run
"""
import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers

# 复用 data-pipeline 模块
_pipeline_root = Path(__file__).resolve().parents[2] / "data-pipeline" / "scripts"
sys.path.insert(0, str(_pipeline_root))

from extractors.minimax_image_extractor import MiniMaxImageExtractor
from transformers.portfolio_excel_transformer import PortfolioExcelTransformer
from transformers.image_portfolio_normalizer import normalize_all
from validators.portfolio_validator import (
    validate_basic_info, validate_nav, validate_position,
    ValidationResult,
)
from loaders.mongodb_loader import PortfolioMongoLoader


def save_excel(df: pd.DataFrame, date_str: str, source_root: Path, base_name: str) -> Path:
    """
    保存 Excel 到 source/smart-money/{date}/image/。
    文件名格式：{base_name}.xlsx

    Args:
        df: DataFrame to save
        date_str: Date string (YYYY-MM-DD)
        source_root: Root directory (source/smart-money)
        base_name: Base filename without extension
    """
    image_dir = source_root / date_str / "image"
    image_dir.mkdir(parents=True, exist_ok=True)

    excel_path = image_dir / f"{base_name}.xlsx"
    counter = 1
    while excel_path.exists():
        excel_path = image_dir / f"{base_name}_{counter:02d}.xlsx"
        counter += 1

    df.to_excel(excel_path, index=False, sheet_name="Sheet1")

    # 应用 Excel 格式：持仓比例列设为百分比格式（保留2位小数）
    # MiniMax 返回 5.77 表示 5.77%，需要除以 100 后 Excel 才能正确显示为 5.77%
    wb = load_workbook(excel_path)
    ws = wb.active

    # 找到持仓比例列的索引（0-based，Excel列从1开始）
    header = [cell.value for cell in ws[1]]
    if "持仓比例" in header:
        col_idx = header.index("持仓比例") + 1  # openpyxl列从1开始
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                cell.value = cell.value / 100  # 转为小数（如 0.0577）
                cell.number_format = '0.00%'   # Excel 百分比格式

    wb.save(excel_path)
    return excel_path


def archive_image(image_path: Path, date_str: str, source_root: Path) -> Path:
    """
    归档原始图片到 source/smart-money/{date}/image/。
    生成带秒级精度的新文件名（避免同一分钟内重名）。

    Args:
        image_path: Original image path
        date_str: Date string (YYYY-MM-DD)
        source_root: Root directory

    Returns:
        New archived image path
    """
    image_dir = source_root / date_str / "image"
    image_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"portfolio_{timestamp}"
    new_image_path = image_dir / f"{base_name}{image_path.suffix}"

    counter = 1
    while new_image_path.exists():
        new_image_path = image_dir / f"{base_name}_{counter:02d}{image_path.suffix}"
        counter += 1

    shutil.copy2(str(image_path), str(new_image_path))
    return new_image_path


async def run_pipeline(
    image_path: str,
    date_str: str,
    source_root: Path,
    dry_run: bool = False,
) -> dict:
    """
    完整流程：图片 → MiniMax OCR → Excel → normalize → validate → MongoDB

    Args:
        image_path: Path to the image file
        date_str: Date string (YYYY-MM-DD)
        source_root: Root directory for source data
        dry_run: If True, only run OCR and save Excel, skip MongoDB
    """
    image_path = Path(image_path)
    if not image_path.exists():
        raise FileNotFoundError(f"Image not found: {image_path}")

    # Step 1: MiniMax OCR Image → DataFrame
    print(f"[Step1] 开始 OCR: {image_path}")
    extractor = MiniMaxImageExtractor()
    records = await extractor.extract(str(image_path))

    if not records:
        raise ValueError("OCR 未提取到数据")

    df = records[0]["df"]
    print(f"[Step1] OCR 完成: {len(df)} rows")

    # Step 1b: 归档原始图片 + 保存 Excel
    archived_image = archive_image(image_path, date_str, source_root)
    base_name = archived_image.stem  # portfolio_YYYYMMDD_HHMMSS
    excel_path = save_excel(df, date_str, source_root, base_name)
    print(f"[Step1] 图片归档 → {archived_image}")
    print(f"[Step1] Excel 保存 → {excel_path} ({len(df)} rows)")

    if dry_run:
        return {
            "image": str(archived_image),
            "excel_path": str(excel_path),
            "rows": len(df),
            "dry_run": True,
        }

    # Step 2: Transform DataFrame → nested JSON
    transformer = PortfolioExcelTransformer()
    nested = transformer.transform(records)
    normalized = normalize_all(nested[0])

    print(f"[Step2] Normalized: basic_info={len(normalized['basic_info'])}, "
          f"nav={len(normalized['nav'])}, position={len(normalized['position'])}")

    # Step 3: Validate
    vr_basic = validate_basic_info(normalized["basic_info"])
    vr_nav = validate_nav(normalized["nav"])
    vr_pos = validate_position(normalized["position"])
    merged = ValidationResult()
    merged.merge(vr_basic)
    merged.merge(vr_nav)
    merged.merge(vr_pos)

    if merged.has_errors:
        print(f"[Step3] Validation errors: {merged.errors}")
        raise ValueError(f"Validation failed: {'; '.join(merged.errors)}")

    if merged.has_warnings:
        for w in merged.warnings:
            print(f"[Step3] Validation warning: {w}")

    # Step 4: MongoDB
    loader = PortfolioMongoLoader()
    result = loader.load_all(normalized)

    print(f"[Step4] MongoDB: basic_info={result['basic_info']}, "
          f"nav={result['nav']}, position={result['position']}")

    return {
        "image": str(archived_image),
        "excel_path": str(excel_path),
        "rows": len(df),
        "basic_info": len(normalized["basic_info"]),
        "nav": len(normalized["nav"]),
        "position": len(normalized["position"]),
        "validation": {
            "valid": merged.valid,
            "errors": merged.errors,
            "warnings": merged.warnings,
        },
        "mongodb": {
            "inserted": result.get("inserted", 0),
            "updated": result.get("updated", 0),
            "skipped": result.get("skipped", 0),
        },
    }


def main():
    parser = argparse.ArgumentParser(description="Image Portfolio Pipeline")
    parser.add_argument("-i", "--image", required=True, help="图片路径")
    parser.add_argument("-d", "--date", required=True, help="日期，如 2026-05-09")
    parser.add_argument("--dry-run", action="store_true", help="仅 OCR 并保存 Excel，不写入 MongoDB")
    args = parser.parse_args()

    # source_root: 向上两级到 workspace-yquant/skills/data，然后拼接 source/smart-money
    source_root = Path(__file__).resolve().parents[2] / "source" / "smart-money"

    import asyncio
    result = asyncio.run(run_pipeline(args.image, args.date, source_root, dry_run=args.dry_run))
    print(f"\n✅ 完成：{result}")


if __name__ == "__main__":
    main()
