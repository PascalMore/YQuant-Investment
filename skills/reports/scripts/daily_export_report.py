#!/usr/bin/env python3
"""
每日导出报告脚本
从 MongoDB 查询持仓和交易数据，生成 Excel 并通过 Telegram 发送

用法: python daily_export_report.py [--env ENV_FILE]
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
from dotenv import load_dotenv
from pymongo import MongoClient
from openpyxl import load_workbook
from openpyxl.styles import numbers

# ============== 配置 ==============
MONGODB_URI = "mongodb://myq:6812345@172.25.240.1:27017/tradingagents"
MONGODB_DB = "tradingagents"

# Excel 列名映射（中英文）
POSITION_COLUMNS = [
    "日期", "产品代码", "产品名称", "资产名称", "Wind代码",
    "持仓比例", "数量", "市值(本币)", "最新净值", "最新规模"
]

TRADE_COLUMNS = [
    "日期", "产品代码", "产品名称", "资产名称", "Wind代码",
    "变化比例", "变化金额", "方向"
]

# 默认 .env 路径
DEFAULT_ENV_PATH = str(Path(__file__).parent.parent / "daily-market-analysis" / ".env")


def load_config(env_path: str = None) -> dict:
    """加载环境变量配置"""
    if env_path is None:
        env_path = DEFAULT_ENV_PATH
    
    if os.path.exists(env_path):
        load_dotenv(env_path)
    
    return {
        "telegram_bot_token": os.getenv("TELEGRAM_BOT_TOKEN", ""),
        "telegram_chat_id": os.getenv("TELEGRAM_CHAT_ID", ""),
        "daily_happy_chat_id": os.getenv("DAILY_HAPPY_GROUP_CHAT_ID", ""),
    }


def connect_mongodb(uri: str, db_name: str):
    """连接 MongoDB"""
    try:
        client = MongoClient(uri, serverSelectionTimeoutMS=5000)
        client.admin.command("ping")
        db = client[db_name]
        return client, db
    except Exception as e:
        print(f"❌ MongoDB 连接失败: {e}")
        sys.exit(1)


def get_latest_date(db, collection_name: str) -> str:
    """获取指定表的最新日期"""
    try:
        collection = db[collection_name]
        result = collection.find_one(sort=[("date", -1)])
        if result and "date" in result:
            return result["date"]
        return None
    except Exception as e:
        print(f"⚠️ 查询 {collection_name} 最新日期失败: {e}")
        return None


def query_positions(db, latest_date: str, min_holding_ratio: float = 0.04) -> pd.DataFrame:
    """查询持仓数据"""
    try:
        collection = db["portfolio_position"]
        query = {
            "date": latest_date,
            "holding_ratio": {"$gte": min_holding_ratio}
        }
        cursor = collection.find(query).sort("holding_ratio", -1)
        
        data = []
        for doc in cursor:
            data.append({
                "日期": doc.get("date", ""),
                "产品代码": doc.get("product_code", ""),
                "产品名称": doc.get("product_name", ""),
                "资产名称": doc.get("asset_name", ""),
                "Wind代码": doc.get("wind_code", ""),
                "持仓比例": doc.get("holding_ratio", 0),
                "数量": doc.get("quantity", 0),
                "市值(本币)": doc.get("market_value", 0),
                "最新净值": doc.get("nav", 0),
                "最新规模": doc.get("aum", 0),
            })
        
        return pd.DataFrame(data)
    except Exception as e:
        print(f"❌ 查询持仓数据失败: {e}")
        return pd.DataFrame()


def query_trades(db, latest_date: str) -> pd.DataFrame:
    """查询交易数据"""
    try:
        collection = db["portfolio_trade"]
        query = {"date": latest_date}
        cursor = collection.find(query)
        
        data = []
        for doc in cursor:
            data.append({
                "日期": doc.get("date", ""),
                "产品代码": doc.get("product_code", ""),
                "产品名称": doc.get("product_name", ""),
                "资产名称": doc.get("asset_name", ""),
                "Wind代码": doc.get("wind_code", ""),
                "变化比例": doc.get("change_ratio", 0),
                "变化金额": doc.get("change_amount", 0),
                "方向": doc.get("direction", ""),
            })
        
        return pd.DataFrame(data)
    except Exception as e:
        print(f"❌ 查询交易数据失败: {e}")
        return pd.DataFrame()


def format_percentage(df: pd.DataFrame, columns: list) -> pd.DataFrame:
    """将小数转换为百分比格式（显示层面）"""
    df_copy = df.copy()
    for col in columns:
        if col in df_copy.columns:
            df_copy[col] = df_copy[col].apply(
                lambda x: f"{x * 100:.2f}%" if isinstance(x, (int, float)) else x
            )
    return df_copy


def create_excel(positions_df: pd.DataFrame, trades_df: pd.DataFrame, 
                 output_path: str, position_date: str) -> bool:
    """生成 Excel 文件"""
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Sheet1: 交易记录
            trades_df.to_excel(writer, sheet_name="业绩前五交易", index=False)
            
            # Sheet2: 持仓记录
            positions_df.to_excel(writer, sheet_name="业绩前五持仓", index=False)
        
        # 格式化百分比列
        wb = load_workbook(output_path)
        
        # 格式化交易表的百分比列
        ws_trades = wb["业绩前五交易"]
        for col_idx, col_name in enumerate(trades_df.columns, 1):
            if col_name == "变化比例":
                for row in range(2, ws_trades.max_row + 1):
                    cell = ws_trades.cell(row=row, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00%'
        
        # 格式化持仓表的百分比列
        ws_positions = wb["业绩前五持仓"]
        for col_idx, col_name in enumerate(positions_df.columns, 1):
            if col_name == "持仓比例":
                for row in range(2, ws_positions.max_row + 1):
                    cell = ws_positions.cell(row=row, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00%'
        
        wb.save(output_path)
        print(f"✅ Excel 已生成: {output_path}")
        return True
        
    except Exception as e:
        print(f"❌ 生成 Excel 失败: {e}")
        return False


def send_telegram_file(token: str, chat_id: str, file_path: str, 
                       caption: str = "") -> bool:
    """通过 Telegram Bot 发送文件"""
    try:
        url = f"https://api.telegram.org/bot{token}/sendDocument"
        
        with open(file_path, "rb") as f:
            files = {"document": f}
            data = {"chat_id": chat_id, "caption": caption}
            response = requests.post(url, data=data, files=files, timeout=30)
        
        result = response.json()
        if result.get("ok"):
            print(f"✅ 文件已发送到 Telegram: {caption}")
            return True
        else:
            print(f"❌ Telegram 发送失败: {result.get('description', 'Unknown error')}")
            return False
            
    except Exception as e:
        print(f"❌ Telegram 请求失败: {e}")
        return False


def get_chat_id_by_name(token: str, group_name: str = "DailyHappyGroup") -> str:
    """通过 Bot API 获取群组 chat_id（如果知道群名）"""
    try:
        # 获取 Bot 所属的群组列表
        url = f"https://api.telegram.org/bot{token}/getUpdates"
        response = requests.get(url, timeout=10)
        result = response.json()
        
        if result.get("ok"):
            updates = result.get("result", [])
            for update in updates:
                chat = update.get("message", {}).get("chat", {})
                if chat.get("type") in ["group", "supergroup"]:
                    if group_name.lower() in chat.get("title", "").lower():
                        return str(chat.get("id", ""))
        
        return None
    except Exception:
        return None


def main():
    parser = argparse.ArgumentParser(description="每日导出持仓和交易报告")
    parser.add_argument("--env", default=DEFAULT_ENV_PATH, help="环境变量文件路径")
    args = parser.parse_args()
    
    print("=" * 50)
    print("📊 每日导出报告脚本")
    print("=" * 50)
    
    # 1. 加载配置
    print("\n[1/6] 加载配置...")
    config = load_config(args.env)
    bot_token = config.get("telegram_bot_token", "")
    
    # 2. 连接 MongoDB
    print("\n[2/6] 连接 MongoDB...")
    client, db = connect_mongodb(MONGODB_URI, MONGODB_DB)
    print(f"   ✅ 已连接到: {MONGODB_DB}")
    
    # 3. 获取持仓表最新日期
    print("\n[3/6] 查询持仓表最新日期...")
    latest_date = get_latest_date(db, "portfolio_position")
    if not latest_date:
        print("❌ 无法获取持仓表最新日期")
        sys.exit(1)
    print(f"   📅 最新日期: {latest_date}")
    
    # 4. 查询数据
    print(f"\n[4/6] 查询数据 (日期: {latest_date})...")
    
    # 查询持仓数据 (holding_ratio >= 0.04)
    positions_df = query_positions(db, latest_date, min_holding_ratio=0.04)
    print(f"   📦 持仓记录: {len(positions_df)} 条")
    
    # 查询交易数据 (使用持仓表的最新日期)
    trades_df = query_trades(db, latest_date)
    print(f"   📦 交易记录: {len(trades_df)} 条")
    
    if positions_df.empty and trades_df.empty:
        print("⚠️ 没有数据可导出")
        sys.exit(0)
    
    client.close()
    
    # 5. 生成 Excel
    print("\n[5/6] 生成 Excel...")
    excel_filename = f"数据_{latest_date}.xlsx"
    excel_path = f"/tmp/{excel_filename}"
    
    if not create_excel(positions_df, trades_df, excel_path, latest_date):
        sys.exit(1)
    
    # 6. 发送 Telegram
    print("\n[6/6] 发送 Telegram...")
    
    # 确定发送目标
    target_chat_id = config.get("daily_happy_chat_id", "")
    
    # 如果没有配置 DailyHappyGroup chat_id，尝试获取或使用占位符
    if not target_chat_id:
        if bot_token:
            # 尝试通过群名查找
            found_id = get_chat_id_by_name(bot_token, "DailyHappyGroup")
            if found_id:
                target_chat_id = found_id
                print(f"   🔍 通过群名找到 chat_id: {target_chat_id}")
            else:
                # 使用 Pascal 个人 chat_id 作为占位符
                target_chat_id = "6805320916"
                print("   ⚠️ 未找到 DailyHappyGroup，使用 Pascal 个人 chat_id 作为占位符")
        else:
            target_chat_id = "6805320916"
            print("   ⚠️ Telegram Bot Token 未配置，使用 Pascal 个人 chat_id")
    
    caption = f"📊 每日报告 {latest_date}"
    
    if bot_token:
        success = send_telegram_file(bot_token, target_chat_id, excel_path, caption)
        if success:
            print("\n✅ 任务完成!")
        else:
            print("\n⚠️ Telegram 发送失败，请检查配置")
    else:
        print("   ⚠️ 未配置 Telegram Bot Token，跳过发送")
        print(f"   📁 Excel 文件已保存: {excel_path}")
    
    print("=" * 50)


if __name__ == "__main__":
    main()