#!/usr/bin/env python3
"""
月度导出报告脚本
将一个月的数据（交易和持仓）合在一个 Excel 中，发送到 Telegram

用法:
    python monthly_export_report.py --year-month 2026-05
    python monthly_export_report.py  # 默认生成上个月
"""

import argparse
import os
import sys
from datetime import datetime, date
from pathlib import Path

import pandas as pd
import requests
from dotenv import load_dotenv
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font

# ============== 配置 ==============
MONGODB_URI = "mongodb://myq:6812345@172.25.240.1:27017/tradingagents"
MONGODB_DB = "tradingagents"

# Excel 列名映射（与 daily_export_report.py 一致）
POSITION_COLUMNS = [
    "日期", "产品代码", "产品名称", "资产名称", "Wind代码",
    "持仓比例", "数量", "市值(本币)", "最新净值", "最新规模"
]

TRADE_COLUMNS = [
    "日期", "产品代码", "产品名称", "资产名称", "Wind代码",
    "变化比例", "变化金额", "方向"
]

DEFAULT_ENV_PATH = str(Path(__file__).parents[4] / "skills" / ".env")


def load_config(env_path: str = None) -> dict:
    if env_path is None:
        env_path = DEFAULT_ENV_PATH
    
    if os.path.exists(env_path):
        load_dotenv(env_path)
    
    return {
        "telegram_bot_token": os.getenv("TELEGRAM_BOT_TOKEN", ""),
        "telegram_chat_id": os.getenv("TELEGRAM_CHAT_ID", ""),
    }


def connect_mongodb(uri: str, db_name: str):
    client = MongoClient(uri, serverSelectionTimeoutMS=5000)
    client.admin.command("ping")
    return client, client[db_name]


def get_product_info_map(db) -> dict:
    """获取产品代码到产品名称的映射"""
    collection = db["portfolio_basic_info"]
    cursor = collection.find()
    return {
        doc["product_code"]: {
            "product_name": doc.get("product_name", ""),
            "latest_nav": doc.get("latest_nav", 0),
            "latest_aum": doc.get("latest_aum", 0),
        }
        for doc in cursor
    }


def query_positions_by_month(db, year_month: str) -> pd.DataFrame:
    """查询指定月份的所有持仓数据（不过滤 holding_ratio）"""
    start_date = f"{year_month}-01"
    year = int(year_month.split("-")[0])
    month = int(year_month.split("-")[1])
    
    if month == 12:
        end_year = year + 1
        end_month = 1
    else:
        end_year = year
        end_month = month + 1
    end_date = f"{end_year}-{end_month:02d}-01"
    
    collection = db["portfolio_position"]
    query = {
        "position_date": {
            "$gte": start_date,
            "$lt": end_date
        }
    }
    cursor = collection.find(query).sort([("position_date", 1), ("product_code", 1)])
    
    product_info = get_product_info_map(db)
    
    data = []
    for doc in cursor:
        pc = doc.get("product_code", "")
        info = product_info.get(pc, {})
        data.append({
            "日期": doc.get("position_date", ""),
            "产品代码": pc,
            "产品名称": info.get("product_name", ""),
            "资产名称": doc.get("asset_name", ""),
            "Wind代码": doc.get("asset_wind_code", ""),
            "持仓比例": doc.get("holding_ratio", 0),
            "数量": doc.get("shares", 0),
            "市值(本币)": doc.get("market_value", 0),
            "最新净值": info.get("latest_nav", 0),
            "最新规模": info.get("latest_aum", 0),
        })
    
    return pd.DataFrame(data)


def query_trades_by_month(db, year_month: str) -> pd.DataFrame:
    """查询指定月份的所有交易数据"""
    start_date = f"{year_month}-01"
    year = int(year_month.split("-")[0])
    month = int(year_month.split("-")[1])
    
    if month == 12:
        end_year = year + 1
        end_month = 1
    else:
        end_year = year
        end_month = month + 1
    end_date = f"{end_year}-{end_month:02d}-01"
    
    collection = db["portfolio_trade"]
    query = {
        "trade_date": {
            "$gte": start_date,
            "$lt": end_date
        }
    }
    cursor = collection.find(query).sort([("trade_date", 1), ("product_code", 1)])
    
    product_info = get_product_info_map(db)
    
    data = []
    for doc in cursor:
        pc = doc.get("product_code", "")
        info = product_info.get(pc, {})
        data.append({
            "日期": doc.get("trade_date", ""),
            "产品代码": pc,
            "产品名称": info.get("product_name", ""),
            "资产名称": doc.get("asset_name", ""),
            "Wind代码": doc.get("asset_wind_code", ""),
            "变化比例": doc.get("change_ratio", 0),
            "变化金额": doc.get("change_amount", 0),
            "方向": doc.get("direction", ""),
        })
    
    return pd.DataFrame(data)


def create_excel(trades_df: pd.DataFrame, positions_df: pd.DataFrame,
                 output_path: str, year_month: str) -> bool:
    """生成 Excel 文件（与 daily_export_report.py 格式一致）"""
    try:
        wb = Workbook()
        
        # ========== Sheet1: 业绩前五交易 ==========
        ws_trades = wb.active
        ws_trades.title = "业绩前五交易"
        
        headers = list(trades_df.columns)
        for col_idx, h in enumerate(headers, 1):
            cell = ws_trades.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
        
        row_idx = 2
        current_product = None
        for _, row in trades_df.iterrows():
            product = row.get("产品代码", "")
            if current_product is not None and product != current_product:
                row_idx += 1  # 组间空行
            current_product = product
            for col_idx, col in enumerate(headers, 1):
                ws_trades.cell(row=row_idx, column=col_idx, value=row.get(col, ""))
            row_idx += 1
        
        # 格式化交易表百分比列
        for col_idx, col_name in enumerate(headers, 1):
            if col_name == "变化比例":
                for r in range(2, ws_trades.max_row + 1):
                    cell = ws_trades.cell(row=r, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00%'
        
        # ========== Sheet2: 业绩前五持仓 ==========
        ws_positions = wb.create_sheet("业绩前五持仓")
        
        pos_headers = list(positions_df.columns)
        for col_idx, h in enumerate(pos_headers, 1):
            cell = ws_positions.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)
        
        for row_idx, (_, row) in enumerate(positions_df.iterrows(), 2):
            for col_idx, col in enumerate(pos_headers, 1):
                ws_positions.cell(row=row_idx, column=col_idx, value=row.get(col, ""))
        
        # 格式化持仓表百分比列
        for col_idx, col_name in enumerate(pos_headers, 1):
            if col_name == "持仓比例":
                for r in range(2, ws_positions.max_row + 1):
                    cell = ws_positions.cell(row=r, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00%'
        
        wb.save(output_path)
        print(f"✅ Excel 已生成: {output_path}")
        return True
        
    except Exception as e:
        print(f"❌ 生成 Excel 失败: {e}")
        return False


def send_telegram_file(token: str, chat_id: str, file_path: str,
                       caption: str = "") -> bool:
    """通过 Telegram Bot 发送文件"""
    try:
        url = f"https://api.telegram.org/bot{token}/sendDocument"
        proxies = None
        if os.getenv("USE_PROXY", "false").lower() == "true":
            proxy_host = os.getenv("PROXY_HOST", "")
            proxy_port = os.getenv("PROXY_PORT", "")
            if proxy_host and proxy_port:
                proxies = {
                    "http": f"http://{proxy_host}:{proxy_port}",
                    "https": f"http://{proxy_host}:{proxy_port}",
                }
        
        with open(file_path, "rb") as f:
            files = {"document": f}
            data = {"chat_id": chat_id, "caption": caption}
            response = requests.post(url, data=data, files=files, timeout=60, proxies=proxies)
        
        result = response.json()
        if result.get("ok"):
            print(f"✅ 文件已发送到 Telegram")
            return True
        else:
            print(f"❌ Telegram 发送失败: {result.get('description', 'Unknown error')}")
            return False
            
    except Exception as e:
        print(f"❌ Telegram 请求失败: {e}")
        return False


def get_previous_month() -> str:
    """获取上个月的 year-month"""
    today = date.today()
    if today.month == 1:
        return f"{today.year - 1}-12"
    else:
        return f"{today.year}-{today.month - 1:02d}"


def main():
    parser = argparse.ArgumentParser(description="月度导出持仓和交易报告")
    parser.add_argument("--year-month", default=get_previous_month(),
                        help="目标月份，格式 YYYY-MM，默认上个月")
    parser.add_argument("--env", default=DEFAULT_ENV_PATH, help="环境变量文件路径")
    parser.add_argument("--test", action="store_true", help="测试模式：发送到 Pascal 个人账号")
    args = parser.parse_args()
    
    year_month = args.year_month
    print("=" * 50)
    print(f"📊 月度导出报告 - {year_month}")
    print("=" * 50)
    
    # 1. 加载配置
    print("\n[1/5] 加载配置...")
    config = load_config(args.env)
    
    # 2. 连接 MongoDB
    print("\n[2/5] 连接 MongoDB...")
    client, db = connect_mongodb(MONGODB_URI, MONGODB_DB)
    print(f"   ✅ 已连接到: {MONGODB_DB}")
    
    # 3. 查询数据
    print(f"\n[3/5] 查询 {year_month} 数据...")
    
    trades_df = query_trades_by_month(db, year_month)
    print(f"   📦 交易记录: {len(trades_df)} 条")
    
    positions_df = query_positions_by_month(db, year_month)
    print(f"   📦 持仓记录: {len(positions_df)} 条")
    
    if trades_df.empty and positions_df.empty:
        print("⚠️ 没有数据可导出")
        sys.exit(0)
    
    client.close()
    
    # 4. 生成 Excel
    print("\n[4/5] 生成 Excel...")
    excel_filename = f"月度报告_{year_month}.xlsx"
    excel_path = f"/tmp/{excel_filename}"
    
    if not create_excel(trades_df, positions_df, excel_path, year_month):
        sys.exit(1)
    
    # 5. 发送 Telegram
    print("\n[5/5] 发送 Telegram...")
    
    target_chat_id = config.get("daily_happy_chat_id", "") or config.get("telegram_chat_id", "6805320916")
    
    if args.test:
        target_chat_id = config.get("telegram_chat_id", "6805320916")
        print("   🧪 测试模式：发送到 Pascal 个人账号")
    else:
        if not target_chat_id or target_chat_id == "6805320916":
            target_chat_id = "-5256445013"  # DailyHappyGroup
        print(f"   📤 发送到 DailyHappyGroup: {target_chat_id}")
    
    bot_token = config.get("telegram_bot_token", "")
    caption = f"📊 月度报告 {year_month}（交易 {len(trades_df)} 条，持仓 {len(positions_df)} 条）"
    
    if bot_token:
        success = send_telegram_file(bot_token, target_chat_id, excel_path, caption)
        if success:
            print(f"\n✅ 任务完成!")
        else:
            print("\n⚠️ Telegram 发送失败")
    else:
        print("   ⚠️ 未配置 Telegram Bot Token，跳过发送")
        print(f"   📁 Excel 文件已保存: {excel_path}")
    
    print("=" * 50)


if __name__ == "__main__":
    main()